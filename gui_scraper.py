import os
import sys
import shutil
import threading
import queue
import time
import random
import urllib.parse
import re
from dataclasses import dataclass
from tkinter import Tk, Text, StringVar, BooleanVar, END, DISABLED, NORMAL, filedialog, messagebox
from tkinter import ttk

os.environ.setdefault("PYTHONIOENCODING", "utf-8")

# ========= Playwright / Excel =========
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import openpyxl
from openpyxl import Workbook

# ========= Constantes =========
EXPECTED_XLSX_NAME = "icpe_details.xlsx"   # fichier de travail à la racine
DEFAULT_SHEET_NAME = "Feuille1"
DEFAULT_JOB_TITLE = "Responsable HSE"

# Colonnes Excel : A=1 (entreprise), B=2 (URL)
COMPANY_COLUMN = 1
RESULT_COLUMN = 2

# Startpage
STARTPAGE_SEARCH_URL = "https://www.startpage.com/do/search?q="
RESULT_BLOCK_SELECTORS = [
    'div.w-gl__result',
    'article[data-testid="result"]',
    'li[class*="result"]',
    'div[class*="result"]'
]
RESULT_LINK_SELECTORS = [
    'a[data-testid="result-title-a"]',
    'a.w-gl__result-title',
    'h3 a',
    'a'
]
RESULT_SNIPPET_SELECTORS = [
    '.w-gl__description',
    'p[class*="snippet"]',
    'div[class*="snippet"] p',
    'p'
]

# Dossier app (gère PyInstaller)
BASE_DIR = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
APP_DIR = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else BASE_DIR

MS_PLAYWRIGHT_DIR = os.path.join(APP_DIR, "ms-playwright")
if os.path.isdir(MS_PLAYWRIGHT_DIR):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = MS_PLAYWRIGHT_DIR

# ========= Utils =========
def strip_non_bmp(s: str) -> str:
    """Supprime les caractères > U+FFFF (emoji, etc.) pour éviter TclError sur vieux Tk."""
    if not isinstance(s, str):
        s = str(s)
    return "".join(ch if ord(ch) <= 0xFFFF else "?" for ch in s)

def human_sleep(a=0.8, b=1.8):
    time.sleep(random.uniform(a, b))

def log_put(log_q: queue.Queue, msg: str):
    try:
        safe = strip_non_bmp(msg)
        log_q.put(safe + ("\n" if not safe.endswith("\n") else ""))
    except Exception:
        pass

def sanitize_filename(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    return name[:120] if len(name) > 120 else name

def ensure_workbook_exists(work_path: str, sheet_name: str):
    """Crée le xlsx minimal si absent (A1='entreprise')."""
    if not os.path.isfile(work_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.cell(row=1, column=COMPANY_COLUMN, value="entreprise")
        wb.save(work_path)
        wb.close()

def accept_cookies_if_any(page, log_q):
    try:
        candidates_role = [
            {"role": "button", "name": "Accept"},
            {"role": "button", "name": "I agree"},
            {"role": "button", "name": "J'accepte"},
            {"role": "button", "name": "Agree"},
            {"role": "button", "name": "OK"},
        ]
        css_candidates = [
            'button#consent-accept', 'button#accept-choices',
            'button[aria-label*="Agree"]', 'button[aria-label*="accept"]',
            'button:has-text("Accept")', 'button:has-text("I agree")'
        ]
        for c in candidates_role:
            try:
                btn = page.get_by_role(c["role"], name=c["name"])
                if btn.is_visible():
                    btn.click(timeout=1500)
                    log_put(log_q, "[cookies] acceptes (role)")
                    human_sleep(0.3, 0.8)
                    return
            except Exception:
                pass
        for sel in css_candidates:
            try:
                loc = page.locator(sel).first
                if loc.is_visible():
                    loc.click(timeout=1500)
                    log_put(log_q, "[cookies] acceptes (css)")
                    human_sleep(0.3, 0.8)
                    return
            except Exception:
                pass
    except Exception:
        pass

def open_startpage_and_search(page, query, log_q):
    url = STARTPAGE_SEARCH_URL + urllib.parse.quote_plus(query)
    page.goto(url, wait_until="domcontentloaded")
    accept_cookies_if_any(page, log_q)
    try:
        page.wait_for_selector(",".join(RESULT_BLOCK_SELECTORS), timeout=4000)
        return
    except PWTimeout:
        pass
    try:
        page.fill('input[name="query"], input[name="q"]', query)
        page.press('input[name="query"], input[name="q"]', "Enter")
    except Exception:
        pass

def extract_results(page):
    for sel in RESULT_BLOCK_SELECTORS:
        blocks = page.query_selector_all(sel)
        if blocks:
            return blocks
    return []

def get_link_and_snippet(block):
    url = None
    title_text = ""
    snippet_text = ""
    for lsel in RESULT_LINK_SELECTORS:
        try:
            link_elem = block.query_selector(lsel)
            if link_elem:
                url = link_elem.get_attribute("href")
                try:
                    title_text = (link_elem.inner_text() or "").strip()
                except Exception:
                    title_text = ""
                break
        except Exception:
            continue
    if not snippet_text:
        for ssel in RESULT_SNIPPET_SELECTORS:
            try:
                sn = block.query_selector(ssel)
                if sn:
                    snippet_text = (sn.inner_text() or "").strip()
                    break
            except Exception:
                continue
    return url, snippet_text, title_text

def search_linkedin_profile(page, company_name, job_title, log_q):
    try:
        query = f'"{job_title}" site:linkedin.com/in "{company_name}"'
        log_put(log_q, f"Recherche: {query}")

        open_startpage_and_search(page, query, log_q)
        try:
            page.wait_for_selector(",".join(RESULT_BLOCK_SELECTORS), timeout=10000)
        except PWTimeout:
            log_put(log_q, "Aucun resultat (timeout)")
            return None

        blocks = extract_results(page)
        log_put(log_q, f"Nombre de resultats: {len(blocks)}")

        company_lc = company_name.lower()
        for block in blocks[:8]:
            url, snippet, title = get_link_and_snippet(block)
            if not url:
                continue
            if "linkedin.com/in/" not in url:
                continue
            snippet_lc = (snippet or "").lower()
            title_lc = (title or "").lower()
            if company_lc in snippet_lc or company_lc in title_lc:
                clean_url = url.split('?')[0]
                log_put(log_q, f"Profil trouve: {clean_url}")
                return clean_url

        log_put(log_q, "Aucun profil correspondant")
        return None

    except Exception as e:
        log_put(log_q, f"Erreur recherche: {e}")
        return None

# ========= Run =========
@dataclass
class RunConfig:
    excel_path: str
    sheet_name: str
    job_title: str
    headless: bool
    fast: bool
    test_mode: bool

def run_scraper(config: RunConfig, stop_event: threading.Event, log_q: queue.Queue, update_q: queue.Queue):
    companies_processed = 0
    urls_found = 0

    log_put(log_q, "Demarrage du scraper (Startpage)")
    log_put(log_q, f"Mode: {'TEST (10 entreprises)' if config.test_mode else 'COMPLET'}")
    log_put(log_q, f"Navigateur: {'Invisible' if config.headless else 'Visible'}")
    log_put(log_q, "-" * 60)

    ensure_workbook_exists(config.excel_path, config.sheet_name)

    try:
        workbook = openpyxl.load_workbook(config.excel_path)
        ws = workbook[config.sheet_name] if config.sheet_name in workbook.sheetnames else workbook.active
    except Exception as e:
        log_put(log_q, f"Erreur ouverture Excel: {e}")
        return

    # A1 et B1
    if (ws.cell(row=1, column=COMPANY_COLUMN).value or "").strip().lower() != "entreprise":
        ws.cell(row=1, column=COMPANY_COLUMN, value="entreprise")
    ws.cell(row=1, column=RESULT_COLUMN, value=config.job_title)

    max_row = ws.max_row
    if config.test_mode:
        max_row = min(11, max_row)

    slow_mo = 0 if (config.fast or config.headless) else 250

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=config.headless,
            slow_mo=slow_mo,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--disable-features=IsolateOrigins,site-per-process'
            ]
        )
        page = browser.new_page(
            viewport={"width": 1400, "height": 900},
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
        )

        try:
            for row in range(2, max_row + 1):
                if stop_event.is_set():
                    log_put(log_q, "Arret demande. Sauvegarde…")
                    break

                company_name = (ws.cell(row=row, column=COMPANY_COLUMN).value or "").strip()
                if not company_name:
                    continue

                companies_processed += 1
                log_put(log_q, f"[{companies_processed}] Entreprise: {company_name}")
                log_put(log_q, "-" * 40)

                linkedin_url = search_linkedin_profile(page, company_name, config.job_title, log_q)
                if linkedin_url:
                    ws.cell(row=row, column=RESULT_COLUMN, value=linkedin_url)
                    urls_found += 1
                    update_q.put({"row": row, "url": linkedin_url})
                    log_put(log_q, f"URL ecrite en B{row}")
                else:
                    log_put(log_q, "Pas de match")

                if companies_processed % 5 == 0:
                    try:
                        workbook.save(config.excel_path)
                        log_put(log_q, f"Sauvegarde intermediaire ({companies_processed})")
                    except Exception as e:
                        log_put(log_q, f"Sauvegarde intermediaire: {e}")

                if row < max_row and not config.fast:
                    base = 1.2 if config.headless else 0.7
                    time.sleep(base + random.uniform(0.2, 0.9))

        except Exception as e:
            log_put(log_q, f"Erreur inattendue: {e}")
        finally:
            try:
                workbook.save(config.excel_path)
                workbook.close()
            except Exception as e:
                log_put(log_q, f"Sauvegarde finale: {e}")
            browser.close()

    log_put(log_q, "\n" + "=" * 60)
    log_put(log_q, "RESUME")
    log_put(log_q, "=" * 60)
    log_put(log_q, f"Entreprises traitees: {companies_processed}")
    log_put(log_q, f"URLs trouvees: {urls_found}")

# ========= Interface (grille) =========
class App:
    def __init__(self, root: Tk):
        self.root = root
        root.title("Scraper LinkedIn (Startpage) - Feuille integree")
        root.geometry("980x640")

        # Vars
        self.excel_path = os.path.join(APP_DIR, EXPECTED_XLSX_NAME)
        self.sheet_name = StringVar(value=DEFAULT_SHEET_NAME)
        self.job_title = StringVar(value=DEFAULT_JOB_TITLE)
        self.headless = BooleanVar(value=False)
        self.fast = BooleanVar(value=False)
        self.test_mode = BooleanVar(value=False)

        self.log_q = queue.Queue()
        self.update_q = queue.Queue()
        self.worker_thread = None
        self.stop_event = threading.Event()

        # dernier dossier d'export (par defaut Home puis fallback APP_DIR)
        self.last_export_dir = os.path.expanduser("~")
        if not os.path.isdir(self.last_export_dir):
            self.last_export_dir = APP_DIR

        ensure_workbook_exists(self.excel_path, self.sheet_name.get())

        # Layout
        wrapper = ttk.Frame(root, padding=10)
        wrapper.pack(fill="both", expand=True)

        # Top controls
        top = ttk.Frame(wrapper)
        top.pack(fill="x", pady=(0,8))

        ttk.Label(top, text=f"Fichier: {os.path.basename(self.excel_path)}  •  Feuille:").pack(side="left")
        self.entry_sheet = ttk.Entry(top, textvariable=self.sheet_name, width=18)
        self.entry_sheet.pack(side="left", padx=(6,18))

        ttk.Label(top, text="Intitule (B1):").pack(side="left")
        self.entry_job = ttk.Entry(top, textvariable=self.job_title, width=28)
        self.entry_job.pack(side="left", padx=(6,18))

        ttk.Checkbutton(top, text="Headless", variable=self.headless).pack(side="left", padx=(0,12))
        ttk.Checkbutton(top, text="Rapide (--fast)", variable=self.fast).pack(side="left", padx=(0,12))
        ttk.Checkbutton(top, text="Mode test (10)", variable=self.test_mode).pack(side="left")

        # Grid controls
        grid_bar = ttk.Frame(wrapper)
        grid_bar.pack(fill="x", pady=(4,6))
        ttk.Button(grid_bar, text="Ajouter ligne", command=self.add_row).pack(side="left")
        ttk.Button(grid_bar, text="Supprimer", command=self.delete_selected).pack(side="left", padx=(6,0))
        ttk.Button(grid_bar, text="Coller entreprises (Ctrl+V)", command=self.paste_lines).pack(side="left", padx=(12,0))
        ttk.Button(grid_bar, text="Exporter la copie", command=self.export_copy).pack(side="right")

        # Treeview (feuille)
        self.tree = ttk.Treeview(wrapper, columns=("entreprise", "url"), show="headings", selectmode="extended")
        self.tree.heading("entreprise", text="entreprise (col A)")
        self.tree.heading("url", text="URL (col B)")
        self.tree.column("entreprise", width=420, anchor="w")
        self.tree.column("url", width=480, anchor="w")
        self.tree.pack(fill="both", expand=True)

        # Inline edit colonne A
        self.tree.bind("<Double-1>", self.begin_edit_cell)
        self.root.bind("<Control-v>", lambda e: self.paste_lines())

        # Bottom controls
        bottom = ttk.Frame(wrapper)
        bottom.pack(fill="x", pady=(8,0))
        self.btn_start = ttk.Button(bottom, text="Lancer", command=self.on_start)
        self.btn_start.pack(side="left")
        self.btn_stop = ttk.Button(bottom, text="Arreter", command=self.on_stop, state=DISABLED)
        self.btn_stop.pack(side="left", padx=(6,0))

        ttk.Label(bottom, text="Journal :").pack(side="left", padx=(18,6))
        self.txt = Text(wrapper, height=10, wrap="word", state=DISABLED)
        self.txt.pack(fill="x", pady=(4,0))

        # Charger la feuille dans la grille
        self.load_sheet_to_grid()

        # Polling
        self.root.after(120, self.flush_queues)

    # --- Grille / Excel ---
    def load_sheet_to_grid(self):
        self.tree.delete(*self.tree.get_children())
        ensure_workbook_exists(self.excel_path, self.sheet_name.get().strip() or DEFAULT_SHEET_NAME)
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb[self.sheet_name.get()] if self.sheet_name.get() in wb.sheetnames else wb.active
            if (ws.cell(row=1, column=COMPANY_COLUMN).value or "").strip().lower() != "entreprise":
                ws.cell(row=1, column=COMPANY_COLUMN, value="entreprise")
            b1 = (ws.cell(row=1, column=RESULT_COLUMN).value or "").strip()
            if b1:
                self.job_title.set(b1)
            max_row = ws.max_row
            for r in range(2, max_row + 1):
                a = (ws.cell(row=r, column=COMPANY_COLUMN).value or "")
                b = (ws.cell(row=r, column=RESULT_COLUMN).value or "")
                if a or b:
                    self.tree.insert("", "end", values=(a, b))
            wb.save(self.excel_path)
            wb.close()
        except Exception as e:
            messagebox.showerror("Excel", f"Lecture feuille impossible:\n{e}")

    def save_grid_to_excel(self):
        ensure_workbook_exists(self.excel_path, self.sheet_name.get().strip() or DEFAULT_SHEET_NAME)
        wb = openpyxl.load_workbook(self.excel_path)
        if self.sheet_name.get() in wb.sheetnames:
            ws = wb[self.sheet_name.get()]
        else:
            ws = wb.active
            ws.title = self.sheet_name.get()

        ws.cell(row=1, column=COMPANY_COLUMN, value="entreprise")
        ws.cell(row=1, column=RESULT_COLUMN, value=self.job_title.get().strip() or DEFAULT_JOB_TITLE)

        ws.delete_rows(2, ws.max_row)
        r = 2
        for iid in self.tree.get_children():
            entreprise, url = self.tree.item(iid, "values")
            ws.cell(row=r, column=COMPANY_COLUMN, value=(entreprise or "").strip())
            ws.cell(row=r, column=RESULT_COLUMN, value=(url or "").strip())
            r += 1

        wb.save(self.excel_path)
        wb.close()

    # --- Actions UI ---
    def add_row(self):
        self.tree.insert("", "end", values=("", ""))

    def delete_selected(self):
        for iid in self.tree.selection():
            self.tree.delete(iid)

    def paste_lines(self):
        try:
            text = self.root.clipboard_get()
        except Exception:
            messagebox.showinfo("Presse-papier", "Rien a coller.")
            return
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            return
        for l in lines:
            self.tree.insert("", "end", values=(l, ""))

    def begin_edit_cell(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        column = self.tree.identify_column(event.x)
        if column != "#1":
            return
        rowid = self.tree.identify_row(event.y)
        if not rowid:
            return
        x, y, w, h = self.tree.bbox(rowid, column)
        value = self.tree.item(rowid, "values")[0]

        entry = ttk.Entry(self.tree)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, value)
        entry.focus()

        def save_edit(e=None):
            new_val = entry.get()
            url = self.tree.item(rowid, "values")[1]
            self.tree.item(rowid, values=(new_val, url))
            entry.destroy()

        entry.bind("<Return>", save_edit)
        entry.bind("<Escape>", lambda e: entry.destroy())
        entry.bind("<FocusOut>", save_edit)

    def export_copy(self):
        """Enregistre une copie a l'emplacement choisi (boite 'Enregistrer sous...')."""
        try:
            self.save_grid_to_excel()
        except Exception as e:
            messagebox.showerror("Excel", f"Sauvegarde Excel impossible:\n{e}")
            return

        title = self.job_title.get().strip() or DEFAULT_JOB_TITLE
        default_name = sanitize_filename(title) + ".xlsx"

        path = filedialog.asksaveasfilename(
            title="Enregistrer la copie sous...",
            initialdir=self.last_export_dir,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel (*.xlsx)", "*.xlsx")]
        )
        if not path:
            return

        rootp, ext = os.path.splitext(path)
        if ext.lower() != ".xlsx":
            path = rootp + ".xlsx"

        try:
            shutil.copy2(self.excel_path, path)
            self.last_export_dir = os.path.dirname(path)
            messagebox.showinfo("Export", f"Copie creee :\n{path}")
        except Exception as e:
            messagebox.showerror("Export", f"Impossible de creer la copie:\n{e}")

    def on_start(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showinfo("En cours", "Une execution est deja en cours.")
            return

        try:
            self.save_grid_to_excel()
        except Exception as e:
            messagebox.showerror("Excel", f"Sauvegarde Excel impossible:\n{e}")
            return

        cfg = RunConfig(
            excel_path=self.excel_path,
            sheet_name=self.sheet_name.get().strip() or DEFAULT_SHEET_NAME,
            job_title=self.job_title.get().strip() or DEFAULT_JOB_TITLE,
            headless=self.headless.get(),
            fast=self.fast.get(),
            test_mode=self.test_mode.get()
        )

        self.stop_event.clear()
        self.btn_start.config(state=DISABLED)
        self.btn_stop.config(state=NORMAL)

        try:
            while True:
                self.log_q.get_nowait()
        except queue.Empty:
            pass
        try:
            while True:
                self.update_q.get_nowait()
        except queue.Empty:
            pass

        self.worker_thread = threading.Thread(
            target=run_scraper, args=(cfg, self.stop_event, self.log_q, self.update_q), daemon=True
        )
        self.worker_thread.start()

    def on_stop(self):
        if self.worker_thread and self.worker_thread.is_alive():
            self.stop_event.set()
            self.append_log("Demande d'arret envoyee...\n")
        else:
            self.append_log("Aucun run en cours.\n")

    # --- Boucle d’UI ---
    def flush_queues(self):
        try:
            while True:
                msg = self.log_q.get_nowait()
                self.append_log(msg)
        except queue.Empty:
            pass

        try:
            while True:
                up = self.update_q.get_nowait()
                row = up.get("row", 0)
                url = up.get("url", "")
                idx = row - 2
                if idx >= 0:
                    children = self.tree.get_children()
                    if idx < len(children):
                        iid = children[idx]
                        entreprise, _ = self.tree.item(iid, "values")
                        self.tree.item(iid, values=(entreprise, url))
        except queue.Empty:
            pass

        if self.worker_thread and not self.worker_thread.is_alive():
            self.btn_start.config(state=NORMAL)
            self.btn_stop.config(state=DISABLED)

        self.root.after(120, self.flush_queues)

    def append_log(self, msg: str):
        safe = strip_non_bmp(msg)
        self.txt.config(state=NORMAL)
        self.txt.insert(END, safe)
        self.txt.see(END)
        self.txt.config(state=DISABLED)

def main():
    root = Tk()
    style = ttk.Style()
    try:
        style.theme_use("clam")
    except Exception:
        pass
    App(root)
    root.mainloop()

if __name__ == "__main__":
    main()
